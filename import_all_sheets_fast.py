#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""快速导入所有工作表"""

import sys
from pathlib import Path
from datetime import datetime
import openpyxl
import csv

# 设置控制台编码
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

PROJECT_ROOT = Path(__file__).parent
EXPORTED_XLSX = PROJECT_ROOT / "exported_from_wechat.xlsx"
PRODUCTS_CSV = PROJECT_ROOT / "assets" / "products.csv"

def convert_value(value):
    if value is None:
        return ''
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    return str(value).strip()

def main():
    print('='*70)
    print('正在导入所有工作表...')
    print('='*70)

    if not EXPORTED_XLSX.exists():
        print(f'文件不存在: {EXPORTED_XLSX}')
        return 1

    try:
        workbook = openpyxl.load_workbook(EXPORTED_XLSX, data_only=True, read_only=True)
        all_data = []
        all_headers = set()

        sheet_names = workbook.sheetnames
        print(f'工作表数量: {len(sheet_names)}')

        for idx, sheet_name in enumerate(sheet_names, 1):
            try:
                sheet = workbook[sheet_name]
                print(f'{idx}/{len(sheet_names)} 处理: {sheet_name}')

                headers = None
                data_count = 0

                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    if row_idx == 1:
                        headers = [str(h) if h else '' for h in row]
                        continue

                    if headers and any(row):
                        row_dict = {}
                        for header, value in zip(headers, row):
                            if header:
                                row_dict[header] = convert_value(value)
                        if row_dict:
                            all_data.append(row_dict)
                            all_headers.add(header)
                            data_count += 1

                print(f'  提取: {data_count} 条记录')

            except Exception as e:
                print(f'  错误: {e}')
                continue

        workbook.close()

        print(f'\n总产品数: {len(all_data)}')
        print(f'总字段数: {len(all_headers)}')

        # 统计航司
        airlines = {}
        for product in all_data:
            name = product.get('产品名称', '') or product.get('产品', '')
            if name:
                code = name[:2] if len(name) >= 2 else '未知'
                airlines[code] = airlines.get(code, 0) + 1

        print(f'\n航司分布:')
        for code, count in sorted(airlines.items()):
            print(f'  {code}: {count}')

        # 保存
        print(f'\n保存到: {PRODUCTS_CSV}')
        final_headers = [h for h in all_headers if h]
        with open(PRODUCTS_CSV, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=final_headers, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(all_data)

        print('完成!')

        return 0

    except Exception as e:
        print(f'错误: {e}')
        import traceback
        traceback.print_exc()
        return 1

if __name__ == '__main__':
    sys.exit(main())
