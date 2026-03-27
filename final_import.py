#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""最终版本的产品导入工具"""

import sys
from pathlib import Path
from datetime import datetime
import openpyxl
import csv

# 设置控制台编码
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

PROJECT_ROOT = Path(__file__).parent
EXPORTED_XLSX = PROJECT_ROOT / "exported_from_wechat.xlsx"
PRODUCTS_CSV = PROJECT_ROOT / "assets" / "products.csv"

def convert_value(value):
    if value is None:
        return ''
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    return str(value).strip()

def main():
    print('='*70)
    print('最终产品数据导入')
    print('='*70)

    if not EXPORTED_XLSX.exists():
        print(f'文件不存在: {EXPORTED_XLSX}')
        return 1

    try:
        workbook = openpyxl.load_workbook(EXPORTED_XLSX, data_only=True, read_only=True)
        all_products = []

        print(f'工作表数量: {len(workbook.sheetnames)}')

        # 标准字段名
        standard_fields = ['产品名称', '航线', '订座舱位', '上浮价格', '政策返点',
                         '产品代码', '出票OFFICE', '备注', '产品有限期']

        for sheet_name in workbook.sheetnames:
            if sheet_name == '产品汇总表25年10.21':
                continue  # 跳过汇总表

            try:
                sheet = workbook[sheet_name]
                print(f'\\n处理: {sheet_name}')

                headers = None
                row_count = 0

                for row in sheet.iter_rows(values_only=True):
                    if not headers:
                        # 第一行是标题
                        headers = [str(h) if h else '' for h in row[:len(standard_fields)]]
                        # 确保有足够的字段
                        if len(headers) < len(standard_fields):
                            headers = standard_fields
                        continue

                    # 提取数据
                    product = {}
                    for i, field in enumerate(standard_fields):
                        value = convert_value(row[i]) if i < len(row) else ''
                        product[field] = value

                    # 跳过空行
                    if any(product.values()):
                        all_products.append(product)
                        row_count += 1

                print(f'  提取: {row_count} 条记录')

            except Exception as e:
                print(f'  错误: {e}')
                continue

        workbook.close()

        print(f'\\n总产品数: {len(all_products)}')

        # 统计航司
        airlines = {}
        for p in all_products:
            name = p.get('产品名称', '')
            if name:
                # 提取航司代码（查找MU、CZ、CA等）
                import re
                match = re.search(r'^([A-Z]{2})', name)
                if match:
                    code = match.group(1)
                elif name.startswith('MU、'):
                    code = 'MU'
                elif name.startswith('CZ'):
                    code = 'CZ'
                else:
                    code = name[:2]

                airlines[code] = airlines.get(code, 0) + 1

        print(f'\\n航司分布 (前20):')
        for code, count in sorted(airlines.items(), key=lambda x: x[1], reverse=True)[:20]:
            print(f'  {code}: {count}')

        # 保存
        print(f'\\n保存到: {PRODUCTS_CSV}')
        with open(PRODUCTS_CSV, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=standard_fields)
            writer.writeheader()
            writer.writerows(all_products)

        print('\\n完成!')
        print(f'总计: {len(all_products)} 条产品')

        return 0

    except Exception as e:
        print(f'错误: {e}')
        import traceback
        traceback.print_exc()
        return 1

if __name__ == '__main__':
    sys.exit(main())
