#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""修复产品数据，正确识别航司"""

import csv
import sys
import shutil
from pathlib import Path
from datetime import datetime

# 设置控制台编码
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

PROJECT_ROOT = Path(__file__).parent
EXPORTED_XLSX = PROJECT_ROOT / "exported_from_wechat.xlsx"
PRODUCTS_CSV = PROJECT_ROOT / "assets" / "products.csv"
BACKUP_DIR = PROJECT_ROOT / "assets" / "backups"

# 创建备份目录
BACKUP_DIR.mkdir(exist_ok=True)

# 备份现有文件
backup_file = BACKUP_DIR / f"products_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
shutil.copy(PRODUCTS_CSV, backup_file)
print(f'✅ 已备份到: {backup_file}')

# 航司映射表
AIRLINE_MAP = {
    'MU': '东航',
    'CZ': '南航',
    'HU': '海南航空',
    'CA': '国航',
    '3U': '四川航空',
    '8L': '祥鹏航空',
    'KY': '昆明航空',
    'ZH': '深圳航空',
    'EU': '成都航空',
    '9H': '长安航空',
    'HO': '吉祥航空',
    'BK': '奥凯航空',
    'GS': '天津航空',
    'G5': '华夏航空',
    'MF': '厦门航空',
    'SC': '山东航空',
    'TV': '西藏航空',
    'GJ': '长龙航空',
    'DR': '桂林航空',
    'QW': '青岛航空',
    'NS': '河北航空',
    'RY': '江西航空',
    'GY': '多彩贵州',
    'DZ': '东海航空',
    'LT': '龙江航空',
    'OQ': '重庆航空',
    'JD': '首都航空',
    'PN': '西部航空',
    'FU': '福州航空',
    'GX': '北部湾航空',
    'UQ': '乌鲁木齐航空',
    'A6': '红土航空',
    'JR': '幸福航空',
    'GP': 'GP',
    '低碳': '低碳'
}

def extract_airline_from_name(product_name):
    """从产品名称中提取航司代码"""
    if not product_name:
        return '未知', '未知'
    
    # 尝试从产品名称前2个字符提取
    if len(product_name) >= 2:
        code = product_name[:2]
        if code in AIRLINE_MAP:
            return code, AIRLINE_MAP[code]
    
    # 如果没有匹配，检查是否包含航司名称
    if '东航' in product_name:
        return 'MU', '东航'
    elif '南航' in product_name:
        return 'CZ', '南航'
    elif '国航' in product_name:
        return 'CA', '国航'
    elif '海南航空' in product_name or '海航' in product_name:
        return 'HU', '海南航空'
    elif '四川航空' in product_name:
        return '3U', '四川航空'
    elif '厦门航空' in product_name:
        return 'MF', '厦门航空'
    elif '山东航空' in product_name:
        return 'SC', '山东航空'
    
    return '未知', product_name[:5]  # 返回前5个字符作为标识

# 读取Excel文件
print('📖 读取Excel文件...')
import openpyxl

wb = openpyxl.load_workbook(EXPORTED_XLSX, data_only=True)
print(f'✅ 找到 {len(wb.sheetnames)} 个工作表')

# 收集所有产品
all_products = []
airlines_count = {}

for sheet_name in wb.sheetnames:
    # 跳过汇总表
    if '汇总' in sheet_name:
        continue
    
    # 从工作表名称获取航司代码
    sheet_code = sheet_name.split()[0] if ' ' in sheet_name else sheet_name
    
    # 特殊处理
    if sheet_name == 'MU':
        sheet_code = 'MU'
        sheet_airline = '东航'
    elif sheet_name == 'HU':
        sheet_code = 'HU'
        sheet_airline = '海南航空'
    elif sheet_name == '8L':
        sheet_code = '8L'
        sheet_airline = '祥鹏航空'
    elif sheet_name == 'A6':
        sheet_code = 'A6'
        sheet_airline = '红土航空'
    elif sheet_name == 'JD':
        sheet_code = 'JD'
        sheet_airline = '首都航空'
    elif sheet_name == 'PN':
        sheet_code = 'PN'
        sheet_airline = '西部航空'
    elif '福州' in sheet_name:
        sheet_code = 'FU'
        sheet_airline = '福州航空'
    elif '北部湾' in sheet_name:
        sheet_code = 'GX'
        sheet_airline = '北部湾航空'
    elif '乌鲁木齐' in sheet_name:
        sheet_code = 'UQ'
        sheet_airline = '乌鲁木齐航空'
    elif '长安' in sheet_name:
        sheet_code = '9H'
        sheet_airline = '长安航空'
    elif '青岛' in sheet_name:
        sheet_code = 'QW'
        sheet_airline = '青岛航空'
    elif '河北' in sheet_name:
        sheet_code = 'NS'
        sheet_airline = '河北航空'
    elif '江西' in sheet_name:
        sheet_code = 'RY'
        sheet_airline = '江西航空'
    elif '多彩' in sheet_name:
        sheet_code = 'GY'
        sheet_airline = '多彩贵州'
    elif '东海' in sheet_name:
        sheet_code = 'DZ'
        sheet_airline = '东海航空'
    elif '龙江' in sheet_name:
        sheet_code = 'LT'
        sheet_airline = '龙江航空'
    elif '重庆' in sheet_name:
        sheet_code = 'OQ'
        sheet_airline = '重庆航空'
    else:
        sheet_airline = AIRLINE_MAP.get(sheet_code, sheet_name)
    
    print(f'  处理工作表: {sheet_name} ({sheet_airline})')
    
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    
    if len(rows) < 2:
        continue
    
    # 获取表头
    headers = [str(cell).strip() for cell in rows[0] if cell is not None]
    
    # 读取数据行
    for row in rows[1:]:
        # 确保行有足够的列
        if len(row) < 2:
            continue
        
        # 补全缺失的列
        row = list(row) + [''] * (10 - len(row))
        
        if not any(row[:10]):
            continue
        
        product = {
            '航司代码': sheet_code,
            '航司名称': sheet_airline,
            '产品名称': row[0] if row[0] else '',
            '航线': row[1] if row[1] else '',
            '订座舱位': row[2] if row[2] else '',
            '上浮价格': str(row[3]) if row[3] else '',
            '政策返点': str(row[4]) if row[4] else '',
            '产品代码': row[5] if row[5] else '',
            '出票OFFICE': row[7] if len(row) > 7 and row[7] else '',
            '备注': row[8] if len(row) > 8 and row[8] else '',
            '产品有限期': row[9] if len(row) > 9 and row[9] else ''
        }
        
        # 跳过空产品
        if not product['产品名称']:
            continue
        
        all_products.append(product)
        
        # 统计航司
        if sheet_airline not in airlines_count:
            airlines_count[sheet_airline] = 0
        airlines_count[sheet_airline] += 1

print(f'\n✅ 共读取 {len(all_products)} 条产品')
print(f'✅ 识别 {len(airlines_count)} 个航司')

# 显示航司统计
print('\n📊 航司分布：')
sorted_airlines = sorted(airlines_count.items(), key=lambda x: x[1], reverse=True)
for i, (airline, count) in enumerate(sorted_airlines, 1):
    print(f'  {i:2d}. {airline:12s}: {count:4d} 条')

# 保存到CSV
print(f'\n💾 保存到 {PRODUCTS_CSV}...')
fieldnames = ['航司代码', '航司名称', '产品名称', '航线', '订座舱位', '上浮价格', '政策返点', '产品代码', '出票OFFICE', '备注', '产品有限期']

with open(PRODUCTS_CSV, 'w', encoding='utf-8-sig', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(all_products)

print('✅ 完成！产品数据已更新')
print(f'\n✅ 总计: {len(all_products)} 条产品, {len(airlines_count)} 个航司')
