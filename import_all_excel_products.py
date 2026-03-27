#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""导入所有工作表的Excel产品数据"""

import sys
import shutil
import os
from pathlib import Path
from datetime import datetime
import openpyxl
import csv

# 设置控制台编码
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

PROJECT_ROOT = Path(__file__).parent
EXPORTED_XLSX = PROJECT_ROOT / "exported_from_wechat.xlsx"
PRODUCTS_CSV = PROJECT_ROOT / "assets" / "products.csv"

def convert_row_value(value):
    """转换单元格值"""
    if value is None:
        return ''
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    return str(value).strip()

def read_all_sheets(file_path):
    """读取所有工作表"""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        all_data = []
        all_headers = set()

        # 遍历所有工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f'  读取工作表: {sheet_name}')

            data = []
            headers = None

            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                # 跳过空行
                if not any(row):
                    continue

                if row_idx == 1:
                    headers = list(row)
                    # 清理headers
                    headers = [h if h else '' for h in headers]
                else:
                    if headers:
                        data.append(dict(zip(headers, row)))

            if data and headers:
                # 清理数据
                cleaned_data = []
                for row in data:
                    cleaned_row = {}
                    for header in headers:
                        if header:
                            key = str(header).strip()
                            value = row.get(header, '')
                            cleaned_row[key] = convert_row_value(value)
                    if cleaned_row:
                        cleaned_data.append(cleaned_row)
                        all_headers.add(str(header).strip())

                all_data.extend(cleaned_data)
                print(f'    提取 {len(cleaned_data)} 条记录')

        workbook.close()
        return all_data, list(all_headers)
    except Exception as e:
        print(f'❌ 读取Excel失败: {e}')
        import traceback
        traceback.print_exc()
        return None, None

def main():
    print('='*70)
    print('📊 导入所有工作表的产品数据')
    print('='*70)

    if not EXPORTED_XLSX.exists():
        print(f'❌ 文件不存在: {EXPORTED_XLSX}')
        return 1

    print(f'📁 文件路径: {EXPORTED_XLSX}')
    print(f'📏 文件大小: {EXPORTED_XLSX.stat().st_size:,} bytes')

    # 读取所有工作表
    print('\n📖 读取Excel文件...')
    all_data, all_headers = read_all_sheets(EXPORTED_XLSX)

    if not all_data:
        print('❌ 无法读取Excel文件')
        return 1

    print(f'\n✅ 成功读取所有工作表')
    print(f'📊 总产品数量: {len(all_data)}')

    # 统计数据
    print(f'\n📈 数据统计:')
    airlines = {}
    routes = set()

    for product in all_data:
        # 尝试多个可能的字段名
        name = (product.get('产品名称', '') or
                product.get('产品', '') or
                product.get('名称', '') or '')

        if name:
            airline = name[:2] if len(name) >= 2 else '未知'
            if airline not in airlines:
                airlines[airline] = 0
            airlines[airline] += 1

        route = (product.get('航线', '') or
                 product.get('航程', '') or '')
        if route:
            routes.add(route[:30])

    print(f'  航空公司数: {len(airlines)}')
    print(f'  航线数: {len(routes)}')

    if airlines:
        print(f'\n✈️  航司分布:')
        for airline, count in sorted(airlines.items()):
            print(f'    {airline}: {count} 个产品')

    # 备份现有文件
    print(f'\n💾 备份现有文件...')
    backup_dir = PROJECT_ROOT / "assets" / "backups"
    backup_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if PRODUCTS_CSV.exists():
        backup_file = backup_dir / f"products_backup_{timestamp}.csv"
        shutil.copy2(PRODUCTS_CSV, backup_file)
        print(f'✅ 已备份到: {backup_file.name}')

    # 保存为CSV
    print(f'\n💾 保存为CSV文件...')
    # 确保有必要的字段
    required_fields = ['产品名称', '航线', '订座舱位', '上浮价格', '政策返点',
                      '产品代码', '出票OFFICE', '备注', '产品有限期']

    # 合并所有字段
    final_headers = list(set(all_headers + required_fields))
    final_headers = [h for h in final_headers if h]  # 移除空字符串

    with open(PRODUCTS_CSV, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=final_headers, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(all_data)

    print(f'✅ 文件已保存: {PRODUCTS_CSV}')
    print(f'✅ 编码: UTF-8 with BOM')
    print(f'✅ 字段数: {len(final_headers)}')

    print(f'\n{"="*70}')
    print('✅ 导入完成！')
    print(f'{"="*70}')

    return 0

if __name__ == '__main__':
    sys.exit(main())
