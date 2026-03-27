#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
航空公司产品智能体 (集成文档系统)
基于清洗后的数据创建产品对象和智能体系统，并集成Markdown文档
支持LangChain框架
"""

import pandas as pd
import os
from utils.airline_names import AIRLINE_NAMES

# 航司简称映射
AIRLINE_ALIASES = {
    '东航': 'MU', '上航': 'FM', '南航': 'CZ',
    '国航': 'CA', '海航': 'HU', '川航': '3U',
    '厦航': 'MF', '山航': 'SC', '深航': 'ZH',
    '昆航': 'KY', '长龙': 'GJ', '祥鹏': '8L',
    '首都': 'JD', '幸福': 'JR', '长安': '9H',
    '天津': 'GS', '北部湾': 'GX', '东海': 'DZ',
    '龙江': 'LT', '西部': 'PN', '西藏': 'TV',
    '成都': 'EU', '华夏': 'G5', '江西': 'RY',
    '青岛': 'QW', '河北': 'NS', '重庆': 'OQ',
    '福州': 'FU', '乌鲁木齐': 'UQ', '联合': 'KN',
    '奥凯': 'BK', '吉祥': 'HO', '大新华': 'CN',
    '红土': 'A6', '桂林': 'DR', '多彩贵州': 'GY'
}


def update_null_airlines(agent):
    """智能更新航司为空的产品"""
    def detect_airline_code(product):
        """智能识别航司代码"""
        product_name = product.product_name or ''
        route = product.route or ''
        office = product.office or ''

        # 方法1: 直接从airline字段获取
        if product.airline:
            return product.airline

        # 方法2: 从产品名称中查找航司二字码
        for code in AIRLINE_NAMES.keys():
            if code in product_name:
                return code

        # 方法3: 从产品名称中查找航司全称
        for name, code in AIRLINE_NAMES.items():
            if name in product_name:
                return code

        # 方法4: 从产品名称中查找航司简称
        for alias, code in AIRLINE_ALIASES.items():
            if alias in product_name:
                return code

        # 方法5: 从路线中查找航司全称或简称
        if route:
            for name, code in AIRLINE_NAMES.items():
                if name in route:
                    return code
            for alias, code in AIRLINE_ALIASES.items():
                if alias in route:
                    return code

        # 方法6: 从Office字段中查找航司二字码
        if office:
            for code in AIRLINE_NAMES.keys():
                if code in office:
                    return code

        # 未识别到航司
        return None

    # 统计并更新
    null_airline_products = [p for p in agent.products if not p.airline or str(p.airline).strip() == '']
    updated_count = 0

    # 按识别出的航司分组
    airline_product_map = {}

    for product in null_airline_products:
        detected_code = detect_airline_code(product)

        if detected_code:
            # 更新产品的airline字段
            product.airline = detected_code
            updated_count += 1

            if detected_code not in airline_product_map:
                airline_product_map[detected_code] = []
            airline_product_map[detected_code].append(product)

    # 输出结果
    if updated_count > 0:
        print(f"\n[OK] 智能识别并更新了 {updated_count} 个产品的航司字段")

        print("[更新详情]:")
        for code in sorted(airline_product_map.keys()):
            count = len(airline_product_map[code])
            name = AIRLINE_NAMES.get(code, code)
            print(f"  {code} ({name}): {count} 个产品")
    else:
        print("\n[INFO] 没有需要更新的航司字段")
import sys
import json
from pathlib import Path
from typing import Annotated, List, Dict, Any, Optional
from langchain.tools import tool, ToolRuntime
from langchain.agents import create_agent
from langchain_openai import ChatOpenAI
from langgraph.graph import MessagesState
from langgraph.graph.message import add_messages
from langchain_core.messages import AnyMessage
from coze_coding_utils.runtime_ctx.context import default_headers, new_context
from storage.memory.memory_saver import get_memory_saver

# 设置UTF-8编码输出
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ============================================================================
# 原有业务逻辑类
# ============================================================================

class AirlineProduct:
    """航空公司产品类"""

    def __init__(self, airline, product_name, route, booking_class,
                 price_increase, rebate_ratio, office, remarks, valid_period,
                 ticket_type='ALL', policy_identifier='', policy_code='', product_id=None):
        """
        初始化产品对象

        Args:
            airline: 航空公司代码
            product_name: 产品名称
            route: 航线
            booking_class: 订座舱位
            price_increase: 上浮价格
            rebate_ratio: 后返比例
            office: 出票OFFICE
            remarks: 备注
            valid_period: 有效期
            ticket_type: 票证类型 (GP/BSP/B2B/MULTI/ALL)
            policy_identifier: 产品政策标识
            policy_code: 产品代码
            product_id: 产品唯一ID
        """
        self.airline = airline
        self.product_name = product_name
        self.route = route
        self.booking_class = booking_class
        self.price_increase = price_increase
        self.rebate_ratio = rebate_ratio
        self.office = office
        self.remarks = remarks
        self.valid_period = valid_period
        self.ticket_type = ticket_type  # 票证类型字段
        self.policy_identifier = policy_identifier  # 产品政策标识字段
        self.policy_code = policy_code  # 产品代码字段
        # 生成唯一ID
        if product_id is None:
            import hashlib
            # 基于产品特征生成唯一ID
            unique_str = f"{airline}_{product_name}_{route}_{booking_class}_{policy_code}_{office}"
            self.product_id = hashlib.md5(unique_str.encode()).hexdigest()[:16]
        else:
            self.product_id = product_id

    def __str__(self):
        """字符串表示"""
        return (f"AirlineProduct(airline={self.airline}, "
                f"product={self.product_name}, "
                f"route={self.route}, "
                f"class={self.booking_class}, "
                f"price={self.price_increase})")

    def __repr__(self):
        return self.__str__()

    def to_dict(self):
        """转换为字典"""
        # 处理nan值，转换为空字符串
        def safe_value(val):
            import math
            if val is None or (isinstance(val, float) and math.isnan(val)):
                return ''
            return val

        return {
            'product_id': self.product_id,  # 添加产品ID
            'airline': self.airline,
            'product_name': self.product_name,
            'route': self.route,
            'booking_class': self.booking_class,
            'price_increase': safe_value(self.price_increase),
            'rebate_ratio': safe_value(self.rebate_ratio),
            'office': self.office,
            'remarks': self.remarks,
            'valid_period': self.valid_period,
            'ticket_type': self.ticket_type,  # 票证类型
            'policy_identifier': self.policy_identifier,  # 产品政策标识
            'policy_code': self.policy_code  # 产品代码
        }


class Documentation:
    """文档管理类"""

    def __init__(self, docs_dir=None):
        """
        初始化文档系统
        """
        if docs_dir is None:
            workspace_path = os.getenv("COZE_WORKSPACE_PATH", "/workspace/projects")
            docs_dir = os.path.join(workspace_path, "assets", "docs")

        self.docs_dir = docs_dir
        self.documents = {}
        self.load_all_documents()

    def load_all_documents(self):
        """加载所有Markdown文档"""
        markdown_files = {
            'api_documentation': 'API_DOCUMENTATION.md',
            'agent_usage': 'AGENT_USAGE.md',
            'project_summary': 'PROJECT_SUMMARY.md',
            'quickstart': 'QUICKSTART.md',
            'readme': 'README.md'
        }

        for key, filename in markdown_files.items():
            file_path = os.path.join(self.docs_dir, filename)
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    self.documents[key] = content
            except FileNotFoundError:
                self.documents[key] = ''
                print(f"[X] 文件未找到: {filename}")
            except Exception as e:
                self.documents[key] = ''
                print(f"[X] 读取失败: {filename} - {e}")

        print(f"[OK] 已加载 {len([v for v in self.documents.values() if v])} 个文档")

    def get_document(self, key):
        """获取指定文档"""
        return self.documents.get(key, '')

    def search_document(self, key, keyword):
        """在指定文档中搜索关键词"""
        doc = self.get_document(key)
        if not doc:
            return []

        lines = doc.split('\n')
        results = []

        for i, line in enumerate(lines, 1):
            if keyword.lower() in line.lower():
                results.append({
                    'line_number': i,
                    'content': line
                })

        return results

    def search_all_documents(self, keyword):
        """在所有文档中搜索关键词"""
        results = {}

        for key in self.documents.keys():
            matches = self.search_document(key, keyword)
            if matches:
                results[key] = matches

        return results

    def display_summary(self):
        """显示文档摘要"""
        print("\n" + "="*80)
        print("文档系统摘要")
        print("="*80)

        for key, content in self.documents.items():
            status = "[OK]" if content else "[X]"
            lines = content.count('\n') + 1 if content else 0
            chars = len(content) if content else 0
            print(f"{status} {key:20s}: {lines:4d} 行, {chars:6d} 字符")

        print("="*80)


class IntelligentAgent:
    """航空公司产品智能体 (集成文档系统)"""

    def __init__(self):
        """初始化智能体"""
        self.products = []
        self.airlines = set()
        self.routes = set()
        self.documentation = None

    def load_documentation(self, docs_dir=None):
        """加载文档系统"""
        self.documentation = Documentation(docs_dir)
        self.documentation.display_summary()

    def add_product(self, product):
        """添加产品到智能体"""
        self.products.append(product)
        self.airlines.add(product.airline)
        self.routes.add(product.route)

    def add_products(self, products_list):
        """批量添加产品"""
        for product in products_list:
            self.add_product(product)

    def get_products_by_airline(self, airline):
        """根据航空公司获取产品列表"""
        return [p for p in self.products if p.airline == airline]

    def get_products_by_route(self, route):
        """根据航线获取产品列表"""
        return [p for p in self.products if p.route == route]

    def get_products_by_booking_class(self, booking_class):
        """根据订座舱位获取产品列表"""
        return [p for p in self.products if p.booking_class == booking_class]

    def search_products(self, keyword, ticket_type=None):
        """
        搜索包含关键词的产品

        Args:
            keyword: 搜索关键词
            ticket_type: 票证类型筛选 (GP/BSP/B2B/MULTI/ALL)

        Returns:
            匹配的产品列表
        """
        keyword = str(keyword).lower()
        results = []
        for product in self.products:
            # 检查关键词匹配（增加policy_code和price_increase字段搜索）
            if (keyword in str(product.product_name).lower() or
                keyword in str(product.route).lower() or
                keyword in str(product.remarks).lower() or
                keyword in str(product.policy_code).lower() or
                keyword in str(product.price_increase).lower()):

                # 检查票证类型
                if ticket_type is None or ticket_type == 'ALL' or product.ticket_type == ticket_type or product.ticket_type == 'ALL':
                    results.append(product)
        return results

    def get_products_by_ticket_type(self, ticket_type):
        """
        根据票证类型获取产品列表

        Args:
            ticket_type: 票证类型 (GP/BSP/B2B/MULTI/ALL)

        Returns:
            该票证类型的产品列表
        """
        if ticket_type == 'ALL':
            return self.products
        return [p for p in self.products if p.ticket_type == ticket_type or p.ticket_type == 'ALL']

    def get_product_by_name(self, product_name):
        """根据产品名称精确获取产品"""
        for product in self.products:
            if product.product_name == product_name:
                return product
        return None

    def get_airlines(self):
        """获取所有航空公司列表"""
        return sorted(list(self.airlines))

    def get_routes(self):
        """获取所有航线列表"""
        return sorted(list(self.routes))

    def get_statistics(self):
        """获取统计信息"""
        stats = {
            'total_products': len(self.products),
            'total_airlines': len(self.airlines),
            'total_routes': len(self.routes),
            'products_by_airline': {},
            'products_by_ticket_type': {}  # 新增票证类型统计
        }

        for airline in self.airlines:
            products = self.get_products_by_airline(airline)
            stats['products_by_airline'][airline] = len(products)

        # 统计各票证类型数量
        ticket_types = ['GP', 'BSP', 'B2B', 'MULTI', 'ALL']
        for ttype in ticket_types:
            stats['products_by_ticket_type'][ttype] = len(self.get_products_by_ticket_type(ttype))

        return stats

    def export_to_dataframe(self):
        """导出为DataFrame"""
        data = [product.to_dict() for product in self.products]
        return pd.DataFrame(data)

    def export_to_csv(self, output_file):
        """导出为CSV文件"""
        df = self.export_to_dataframe()
        df.to_csv(output_file, index=False, encoding='utf-8-sig')
        print(f"[OK] 产品数据已导出到: {output_file}")

    def export_to_excel(self, output_file):
        """导出为Excel文件"""
        df = self.export_to_dataframe()
        df.to_excel(output_file, index=False, engine='openpyxl')
        print(f"[OK] 产品数据已导出到: {output_file}")

    def display_summary(self):
        """显示摘要信息"""
        stats = self.get_statistics()
        print("\n" + "="*80)
        print("智能体产品摘要")
        print("="*80)
        print(f"总产品数: {stats['total_products']}")
        print(f"航空公司数: {stats['total_airlines']}")
        print(f"航线数: {stats['total_routes']}")
        print("\n各航空公司产品数量:")
        for airline, count in sorted(stats['products_by_airline'].items()):
            print(f"  {airline}: {count} 个产品")
        print("="*80)


# ============================================================================
# 数据加载函数
# ============================================================================

# ============================================================================
# 辅助函数
# ============================================================================

def extract_ticket_type_and_policy(product_name, route, rebate_ratio):
    """
    从产品名称、航线和政策返点中提取票证类型和产品政策标识

    Args:
        product_name: 产品名称
        route: 航线
        rebate_ratio: 政策返点

    Returns:
        tuple: (票证类型, 产品政策标识)
    """
    import re
    
    ticket_type = 'ALL'
    policy_identifier = ''
    
    # 从产品名称中提取票证类型
    product_name_upper = str(product_name).upper()
    route_upper = str(route).upper()
    rebate_upper = str(rebate_ratio).upper()
    
    # 检测GP (集团客户)
    if 'GP' in product_name_upper or 'GP' in route_upper:
        ticket_type = 'GP'
        # 提取GP相关标识
        gp_match = re.search(r'GP[：:、,，]?\s*([^\s,，]+)', product_name_upper)
        if gp_match:
            policy_identifier = gp_match.group(1)
    
    # 检测BSP (中性票)
    elif 'BSP' in product_name_upper:
        ticket_type = 'BSP'
        # 提取BSP相关标识
        bsp_match = re.search(r'BSP[：:、,，]?\s*([^\s,，]+)', product_name_upper)
        if bsp_match:
            policy_identifier = bsp_match.group(1)
    
    # 检测B2B (企业)
    elif 'B2B' in product_name_upper:
        ticket_type = 'B2B'
        # 提取B2B相关标识
        b2b_match = re.search(r'B2B[：:、,，]?\s*([^\s,，]+)', product_name_upper)
        if b2b_match:
            policy_identifier = b2b_match.group(1)
    
    # 从政策返点中提取产品标识
    if not policy_identifier and rebate_ratio:
        # 提取类似 "L5 300里程" 中的等级和里程
        level_match = re.search(r'L([1-7])', rebate_upper)
        if level_match:
            level = level_match.group(1)
            mileage_match = re.search(r'(\d+)\s*里程', rebate_upper)
            if mileage_match:
                mileage = mileage_match.group(1)
                policy_identifier = f'L{level}-{mileage}里程'
    
    # 从产品名称中提取其他关键标识
    if not policy_identifier:
        # 提取免费服务标识
        if '免费' in product_name:
            if '快速安检' in product_name:
                policy_identifier = '免费安检'
            elif '接送机' in product_name:
                policy_identifier = '免费接送'
            elif 'WiFi' in product_name or 'WIFI' in product_name:
                policy_identifier = '免费WiFi'
            else:
                policy_identifier = '免费服务'
        # 提取里程标识
        elif '里程' in product_name:
            mileage_match = re.search(r'(\d+)\s*里程', product_name_upper)
            if mileage_match:
                policy_identifier = f'{mileage_match.group(1)}里程'
            else:
                policy_identifier = '里程奖励'
        # 提取其他标识
        else:
            # 提取产品关键特征
            keywords = ['优享', '安心', '空铁', '接送', 'WiFi', 'WIFI', '国博', '低碳']
            for keyword in keywords:
                if keyword in product_name:
                    policy_identifier = keyword
                    break
    
    return ticket_type, policy_identifier


# ============================================================================
# 数据加载函数
# ============================================================================

def load_cleaned_data(cleaned_dir=None, use_all_data=True):
    """
    加载清洗后的数据

    Args:
        cleaned_dir (str): 清洗数据目录
        use_all_data (bool): 是否使用合并后的所有数据文件

    Returns:
        DataFrame: 合并后的数据
    """
    if cleaned_dir is None:
        # 获取脚本所在目录的父目录作为workspace路径
        workspace_path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    # 优先从CSV文件加载（用户提供的最新数据）
    if use_all_data:
        # 优先加载最新的products.csv文件（包含最新的9H等航司数据）
        products_file = os.path.join(workspace_path, "assets", "products.csv")
        if os.path.exists(products_file):
            try:
                df = pd.read_csv(products_file, encoding='utf-8-sig')
                # 填充所有nan值为空字符串
                df = df.fillna('')
                # 重命名airline列为airline_code以保持兼容性
                if 'airline' in df.columns and 'airline_code' not in df.columns:
                    df = df.rename(columns={'airline': 'airline_code'})
                # 确保有必要的列（支持中文列名）
                has_english_columns = 'airline_code' in df.columns and 'product_name' in df.columns
                has_chinese_columns = '产品名称' in df.columns

                if has_english_columns:
                    print(f"\n[OK] 从products.csv加载（英文列名）: {len(df)} 行数据")
                    return df
                elif has_chinese_columns:
                    # 中文列名格式，直接返回，create_airline_products会处理
                    print(f"\n[OK] 从products.csv加载（中文列名）: {len(df)} 行数据")
                    return df
                else:
                    print(f"[WARN] products.csv缺少必要的列")
            except Exception as e:
                print(f"[WARN] 加载products.csv失败: {e}")

        # 回退到修复后的汇总CSV文件
        all_data_file = os.path.join(workspace_path, "assets", "airline_products_all_fixed.csv")
        if os.path.exists(all_data_file):
            try:
                df = pd.read_csv(all_data_file, encoding='utf-8-sig')
                # 检查是否有airline_code列
                if 'airline_code' in df.columns:
                    print(f"\n[OK] 从合并CSV文件加载: {len(df)} 行数据")
                    return df
                else:
                    print(f"[WARN] CSV文件缺少airline_code列")
            except Exception as e:
                print(f"[WARN] 加载合并CSV文件失败: {e}")

        # 回退到Excel文件加载（包含31个航司工作表，共885个产品）
        # 首先尝试加载新的汇总表文件
        new_excel_file = os.path.join(workspace_path, "assets", "各航司汇总产品政策更新表_格式应用后.xlsx")
        if os.path.exists(new_excel_file):
            try:
                import openpyxl
                all_data = []
                wb = openpyxl.load_workbook(new_excel_file, read_only=True)
                
                # 读取所有工作表（跳过汇总表）
                for sheet_name in wb.sheetnames:
                    if '汇总' not in sheet_name:
                        try:
                            df = pd.read_excel(new_excel_file, sheet_name=sheet_name)
                            # 添加航司代码列
                            df['airline_code'] = sheet_name
                            all_data.append(df)
                        except Exception as e:
                            print(f"[WARN] 加载工作表 {sheet_name} 失败: {e}")
                
                if all_data:
                    combined_df = pd.concat(all_data, ignore_index=True)
                    print(f"\n[OK] 从新的汇总表加载: {len(combined_df)} 行数据（来自{len(all_data)}个工作表）")
                    return combined_df
            except Exception as e:
                print(f"[WARN] 加载新的汇总表失败: {e}")

        # 回退到旧的Excel文件（包含31个航司工作表，共885个产品）
        excel_file = os.path.join(workspace_path, "assets", "agent_data_with_docs", "所有产品数据_export.xlsx")
        if os.path.exists(excel_file):
            try:
                import openpyxl
                all_data = []
                wb = openpyxl.load_workbook(excel_file, read_only=True)
                
                # 读取所有工作表（跳过汇总表）
                for sheet_name in wb.sheetnames:
                    if '汇总' not in sheet_name:
                        try:
                            df = pd.read_excel(excel_file, sheet_name=sheet_name)
                            # 添加航司代码列
                            df['airline_code'] = sheet_name
                            all_data.append(df)
                        except Exception as e:
                            print(f"[WARN] 加载工作表 {sheet_name} 失败: {e}")
                
                if all_data:
                    combined_df = pd.concat(all_data, ignore_index=True)
                    print(f"\n[OK] 从Excel文件加载: {len(combined_df)} 行数据（来自{len(all_data)}个工作表）")
                    return combined_df
            except ImportError:
                print("[WARN] openpyxl未安装，无法加载Excel文件")
            except Exception as e:
                print(f"[WARN] 加载Excel文件失败: {e}")

        # 回退到带票证类型的CSV文件
        ticket_type_data_file = os.path.join(workspace_path, "assets", "airline_products_all_with_ticket_type.csv")
        if os.path.exists(ticket_type_data_file):
            try:
                df = pd.read_csv(ticket_type_data_file, encoding='utf-8-sig')
                print(f"\n[OK] 从带票证类型的文件加载: {len(df)} 行数据")
                return df
            except Exception as e:
                print(f"[WARN] 加载带票证类型文件失败: {e}")

        # 回退到普通数据文件
        all_data_file = os.path.join(workspace_path, "assets", "airline_products_all.csv")
        if os.path.exists(all_data_file):
            try:
                df = pd.read_csv(all_data_file, encoding='utf-8-sig')
                print(f"\n[OK] 从合并文件加载: {len(df)} 行数据")
                return df
            except Exception as e:
                print(f"[X] 加载合并文件失败: {e}")

    # 回退到旧的加载方式
    if cleaned_dir is None:
        workspace_path = os.getenv("COZE_WORKSPACE_PATH", "/workspace/projects")
        cleaned_dir = os.path.join(workspace_path, "assets", "cleaned_data")

    all_data = []

    # 读取所有清洗后的CSV文件
    csv_files = Path(cleaned_dir).glob('*_cleaned.csv')
    for csv_file in csv_files:
        airline = csv_file.stem.replace('_cleaned', '')
        try:
            df = pd.read_csv(csv_file, encoding='utf-8-sig')
            df['航空公司'] = airline
            all_data.append(df)
        except Exception as e:
            print(f"[X] 加载失败: {airline} - {e}")

    if all_data:
        combined_df = pd.concat(all_data, ignore_index=True)
        print(f"\n总共加载: {len(combined_df)} 行数据")
        return combined_df
    else:
        print("没有加载到任何数据")
        return pd.DataFrame()


def create_airline_products(df):
    """
    从DataFrame创建产品对象列表

    Args:
        df (DataFrame): 数据框

    Returns:
        list: AirlineProduct对象列表
    """
    products = []

    # 检测数据格式
    # 新汇总表格式：包含中文列名，但无airline_code列，需从产品名称中提取
    # 新格式：包含airline_code列（Excel文件）
    # 英文格式：包含airline_code和product_name列（CSV文件）
    # 中文格式：包含航空公司和产品名称列（旧Excel文件）
    # 混合格式：包含中文列名和airline_code列（更新的汇总表）
    
    use_mixed_format = '产品名称' in df.columns and 'airline_code' in df.columns
    use_new_summary_format = '产品名称' in df.columns and 'airline_code' not in df.columns
    use_new_format = 'airline_code' in df.columns
    use_english_format = 'product_name' in df.columns
    use_chinese_format = '产品名称' in df.columns
    use_chinese_with_airline_code = '产品名称' in df.columns and '航司代码' in df.columns

    if use_chinese_with_airline_code:
        # 使用中文列名且包含航司代码字段(新增:直接从CSV的航司代码字段获取)
        for index, row in df.iterrows():
            try:
                # 优先使用"航司代码"字段
                airline = row.get('航司代码', '')
                product_name = row.get('产品名称', '')
                route = row.get('航线', '')
                booking_class = row.get('订座舱位', '')
                price_increase = row.get('上浮价格', 0)
                rebate_ratio = row.get('政策返点', '')
                office = row.get('出票OFFICE', '')
                remarks = row.get('备注', '')
                valid_period = row.get('产品有限期', '')
                policy_code_val = row.get('产品代码', '')
                policy_code = str(policy_code_val).strip() if pd.notna(policy_code_val) else ''

                # 清洗数据
                product_name = str(product_name).strip() if pd.notna(product_name) else ''
                route = str(route).strip() if pd.notna(route) else ''
                booking_class = str(booking_class).strip() if pd.notna(booking_class) else ''
                rebate_ratio = str(rebate_ratio).strip() if pd.notna(rebate_ratio) else ''
                office = str(office).strip() if pd.notna(office) else ''
                remarks = str(remarks).strip() if pd.notna(remarks) else ''
                valid_period = str(valid_period).strip() if pd.notna(valid_period) else ''
                airline = str(airline).strip() if pd.notna(airline) else ''

                # 转换价格为整数
                if pd.notna(price_increase):
                    try:
                        import re
                        price_match = re.search(r'\d+', str(price_increase))
                        if price_match:
                            price_increase = int(price_match.group())
                        else:
                            price_increase = 0
                    except:
                        price_increase = 0
                else:
                    price_increase = 0

                # 提取票证类型和产品政策标识
                ticket_type, policy_identifier = extract_ticket_type_and_policy(
                    str(product_name), str(route), str(rebate_ratio)
                )

                # 只添加有产品名称和航空公司代码的数据
                if product_name and product_name != 'nan' and airline and airline != 'nan':
                    product = AirlineProduct(
                        airline=airline,
                        product_name=product_name,
                        route=route,
                        booking_class=booking_class,
                        price_increase=price_increase,
                        rebate_ratio=rebate_ratio,
                        office=office,
                        remarks=remarks,
                        valid_period=valid_period,
                        ticket_type=ticket_type,
                        policy_identifier=policy_identifier,
                        policy_code=policy_code
                    )
                    products.append(product)

            except Exception as e:
                print(f"  [X] 第 {index} 行处理失败: {e}")
                continue

    elif use_mixed_format:
        # 使用混合格式（中文列名 + airline_code列）
        for index, row in df.iterrows():
            try:
                # 优先使用"航司代码"字段，如果没有则使用"airline_code"
                airline = row.get('航司代码', row.get('airline_code', ''))
                product_name = row.get('产品名称', '')
                route = row.get('航线', '')
                booking_class = row.get('订座舱位', '')
                price_increase = row.get('上浮价格', 0)
                rebate_ratio = row.get('政策返点', '')
                office = row.get('出票OFFICE', '')
                remarks = row.get('备注', '')
                valid_period = row.get('产品有限期', '')
                policy_code_val = row.get('产品代码', '')
                policy_code = str(policy_code_val).strip() if pd.notna(policy_code_val) else ''

                # 清洗数据
                product_name = str(product_name).strip() if pd.notna(product_name) else ''
                route = str(route).strip() if pd.notna(route) else ''
                booking_class = str(booking_class).strip() if pd.notna(booking_class) else ''
                rebate_ratio = str(rebate_ratio).strip() if pd.notna(rebate_ratio) else ''
                office = str(office).strip() if pd.notna(office) else ''
                remarks = str(remarks).strip() if pd.notna(remarks) else ''
                valid_period = str(valid_period).strip() if pd.notna(valid_period) else ''
                airline = str(airline).strip() if pd.notna(airline) else ''

                # 转换价格为整数
                if pd.notna(price_increase):
                    try:
                        import re
                        price_match = re.search(r'\d+', str(price_increase))
                        if price_match:
                            price_increase = int(price_match.group())
                        else:
                            price_increase = 0
                    except:
                        price_increase = 0
                else:
                    price_increase = 0

                # 提取票证类型和产品政策标识
                ticket_type, policy_identifier = extract_ticket_type_and_policy(
                    str(product_name), str(route), str(rebate_ratio)
                )

                # 只添加有产品名称和航空公司代码的数据
                if product_name and product_name != 'nan' and airline and airline != 'nan':
                    product = AirlineProduct(
                        airline=airline,
                        product_name=product_name,
                        route=route,
                        booking_class=booking_class,
                        price_increase=price_increase,
                        rebate_ratio=rebate_ratio,
                        office=office,
                        remarks=remarks,
                        valid_period=valid_period,
                        ticket_type=ticket_type,
                        policy_identifier=policy_identifier,
                        policy_code=policy_code
                    )
                    products.append(product)

            except Exception as e:
                print(f"  [X] 第 {index} 行处理失败: {e}")
                continue
    
    elif use_new_format and use_english_format:
        # 使用新的英文数据格式（从各航空公司Sheet提取或CSV文件）
        for index, row in df.iterrows():
            try:
                airline = row.get('airline_code', '')
                product_name = row.get('product_name', '')
                route = row.get('route', '')
                booking_class = row.get('booking_class', '')
                # 修正：支持'price'和'price_increase'两种列名
                price_increase = row.get('price_increase', row.get('price', 0))
                # 修正：支持'rebate_ratio'和'rebate'两种列名
                rebate_ratio = row.get('rebate_ratio', row.get('rebate', ''))
                office = row.get('office', '')
                remarks = row.get('remarks', '')
                valid_period = row.get('valid_period', row.get('validity', ''))
                policy_code = row.get('policy_code', '')
                policy_identifier = row.get('policy_identifier', '')
                ticket_type = row.get('ticket_type', 'ALL')

                # 清洗数据
                product_name = str(product_name).strip() if pd.notna(product_name) else ''
                route = str(route).strip() if pd.notna(route) else ''
                booking_class = str(booking_class).strip() if pd.notna(booking_class) else ''
                rebate_ratio = str(rebate_ratio).strip() if pd.notna(rebate_ratio) else ''
                office = str(office).strip() if pd.notna(office) else ''
                remarks = str(remarks).strip() if pd.notna(remarks) else ''
                valid_period = str(valid_period).strip() if pd.notna(valid_period) else ''
                airline = str(airline).strip() if pd.notna(airline) else ''
                policy_code = str(policy_code).strip() if pd.notna(policy_code) else ''
                policy_identifier = str(policy_identifier).strip() if pd.notna(policy_identifier) else ''

                # 转换价格为整数
                if pd.notna(price_increase):
                    try:
                        # 如果是字符串，提取其中的数字
                        if isinstance(price_increase, str):
                            import re
                            price_match = re.search(r'\d+', str(price_increase))
                            if price_match:
                                price_increase = int(price_match.group())
                            else:
                                price_increase = 0
                        else:
                            # 如果已经是数字，直接转换
                            price_increase = int(price_increase) if price_increase else 0
                    except:
                        price_increase = 0
                else:
                    price_increase = 0

                # 提取票证类型和产品政策标识
                if not policy_identifier or policy_identifier == '':
                    ticket_type, policy_identifier = extract_ticket_type_and_policy(
                        str(product_name), str(route), str(rebate_ratio)
                    )

                # 只添加有产品名称和航空公司代码的数据
                if product_name and product_name != 'nan' and airline and airline != 'nan':
                    product = AirlineProduct(
                        airline=airline,
                        product_name=product_name,
                        route=route,
                        booking_class=booking_class,
                        price_increase=price_increase,
                        rebate_ratio=rebate_ratio,
                        office=office,
                        remarks=remarks,
                        valid_period=valid_period,
                        ticket_type=ticket_type,
                        policy_identifier=policy_identifier,
                        policy_code=policy_code
                    )
                    products.append(product)

            except Exception as e:
                print(f"  [X] 第 {index} 行处理失败: {e}")
                continue
    
    elif use_new_summary_format:
        # 使用新的汇总表格式（从产品名称中提取航空公司代码）
        for index, row in df.iterrows():
            try:
                product_name = row.get('产品名称', '')
                route = row.get('航线', '')
                booking_class = row.get('订座舱位', '')
                price_increase = row.get('上浮价格', 0)
                rebate_ratio = row.get('政策返点', '')
                office = row.get('出票OFFICE', '')
                remarks = row.get('备注', '')
                valid_period = row.get('产品有限期', '')
                policy_code_val = row['产品代码'] if '产品代码' in row.index else ''
                policy_code = str(policy_code_val) if pd.notna(policy_code_val) else ''

                # 从产品名称中提取航空公司代码（如从 "MU、GP免费快速安检通道" 中提取 "MU"）
                import re

                # 尝试提取航空公司代码（2个大写字母开头，后面跟着中文逗号或顿号）
                airline_match = re.match(r'^([A-Z]{2})[、,，]', str(product_name))
                if not airline_match:
                    # 如果没有找到，尝试简单的2个大写字母开头
                    airline_match = re.match(r'^([A-Z]{2})', str(product_name))

                airline = airline_match.group(1) if airline_match else ''

                # 清洗数据
                product_name = str(product_name).strip() if pd.notna(product_name) else ''
                route = str(route).strip() if pd.notna(route) else ''
                booking_class = str(booking_class).strip() if pd.notna(booking_class) else ''
                rebate_ratio = str(rebate_ratio).strip() if pd.notna(rebate_ratio) else ''
                office = str(office).strip() if pd.notna(office) else ''
                remarks = str(remarks).strip() if pd.notna(remarks) else ''
                valid_period = str(valid_period).strip() if pd.notna(valid_period) else ''
                policy_code = str(policy_code).strip() if pd.notna(policy_code) else ''

                # 转换价格为整数
                if pd.notna(price_increase):
                    try:
                        # 提取价格中的数字
                        price_match = re.search(r'\d+', str(price_increase))
                        if price_match:
                            price_increase = int(price_match.group())
                        else:
                            price_increase = 0
                    except:
                        price_increase = 0
                else:
                    price_increase = 0

                # 提取票证类型和产品政策标识
                ticket_type, policy_identifier = extract_ticket_type_and_policy(
                    str(product_name), str(route), str(rebate_ratio)
                )

                # 获取票证类型（如果存在）
                ticket_type = row.get('ticket_type', 'ALL') if 'ticket_type' in df.columns else ticket_type
                ticket_type = str(ticket_type).strip() if pd.notna(ticket_type) else ticket_type

                # 只添加有产品名称的数据
                if product_name and product_name != 'nan':
                    product = AirlineProduct(
                        airline=airline,
                        product_name=product_name,
                        route=route,
                        booking_class=booking_class,
                        price_increase=price_increase,
                        rebate_ratio=rebate_ratio,
                        office=office,
                        remarks=remarks,
                        valid_period=valid_period,
                        ticket_type=ticket_type,
                        policy_identifier=policy_identifier,
                        policy_code=policy_code
                    )
                    products.append(product)

            except Exception as e:
                print(f"  [X] 第 {index} 行处理失败: {e}")
                continue

    elif use_new_format and use_chinese_format:
        # 使用新的中文数据格式（从Excel文件读取）
        for index, row in df.iterrows():
            try:
                # 优先使用"航司代码"字段，如果没有则使用"airline_code"
                airline = row.get('航司代码', row.get('airline_code', ''))
                product_name = row.get('产品名称', '')
                route = row.get('航线', '')
                booking_class = row.get('订座舱位', '')
                price_increase = row.get('上浮价格', 0)
                rebate_ratio = row.get('政策返点', '')  # 后返信息
                office = row.get('出票OFFICE', '')
                remarks = row.get('备注', '')
                valid_period = row.get('产品有限期', '')

                # 清洗数据
                product_name = str(product_name).strip() if pd.notna(product_name) else ''
                route = str(route).strip() if pd.notna(route) else ''
                booking_class = str(booking_class).strip() if pd.notna(booking_class) else ''
                rebate_ratio = str(rebate_ratio).strip() if pd.notna(rebate_ratio) else ''
                office = str(office).strip() if pd.notna(office) else ''
                remarks = str(remarks).strip() if pd.notna(remarks) else ''
                valid_period = str(valid_period).strip() if pd.notna(valid_period) else ''

                # 转换价格为整数
                if pd.notna(price_increase):
                    try:
                        # 提取价格中的数字
                        import re
                        price_match = re.search(r'\d+', str(price_increase))
                        if price_match:
                            price_increase = int(price_match.group())
                        else:
                            price_increase = 0
                    except:
                        price_increase = 0
                else:
                    price_increase = 0

                # 提取票证类型和产品政策标识
                ticket_type, policy_identifier = extract_ticket_type_and_policy(
                    str(product_name), str(route), str(rebate_ratio)
                )

                # 获取票证类型（如果存在）
                ticket_type = row.get('ticket_type', 'ALL') if 'ticket_type' in df.columns else ticket_type
                ticket_type = str(ticket_type).strip() if pd.notna(ticket_type) else ticket_type

                # 只添加有产品名称的数据
                if product_name and product_name != 'nan':
                    product = AirlineProduct(
                        airline=airline,
                        product_name=product_name,
                        route=route,
                        booking_class=booking_class,
                        price_increase=price_increase,
                        rebate_ratio=rebate_ratio,
                        office=office,
                        remarks=remarks,
                        valid_period=valid_period,
                        ticket_type=ticket_type,
                        policy_identifier=policy_identifier
                    )
                    products.append(product)

            except Exception as e:
                print(f"  [X] 第 {index} 行处理失败: {e}")
                continue
    
    else:
        # 使用旧的数据格式
        for index, row in df.iterrows():
            try:
                # 获取各列的值
                airline = row.get('航空公司', '')
                product_name = row.get('产品名称', row.get('Unnamed: 1', ''))
                route = row.get('航线', row.get('Unnamed: 2', ''))
                booking_class = row.get('订座舱位', row.get('Unnamed: 3', ''))
                price_increase = row.get('上浮价格', row.get('Unnamed: 4', 0))
                rebate_ratio = row.get('做单', row.get('Unnamed: 5', ''))
                office = row.get('出票OFFICE', row.get('Unnamed: 6', ''))
                remarks = row.get('备注', row.get('Unnamed: 7', ''))
                valid_period = row.get('产品有限期', row.get('Unnamed: 8', ''))

                # 清洗数据
                product_name = str(product_name).strip() if pd.notna(product_name) else ''
                route = str(route).strip() if pd.notna(route) else ''
                booking_class = str(booking_class).strip() if pd.notna(booking_class) else ''
                rebate_ratio = str(rebate_ratio).strip() if pd.notna(rebate_ratio) else ''
                office = str(office).strip() if pd.notna(office) else ''
                remarks = str(remarks).strip() if pd.notna(remarks) else ''
                valid_period = str(valid_period).strip() if pd.notna(valid_period) else ''

                # 转换价格为整数
                if pd.notna(price_increase):
                    try:
                        price_increase = int(float(str(price_increase)))
                    except:
                        price_increase = 0
                else:
                    price_increase = 0

                # 提取票证类型和产品政策标识
                ticket_type, policy_identifier = extract_ticket_type_and_policy(
                    str(product_name), str(route), str(rebate_ratio)
                )

                # 获取票证类型（如果存在）
                ticket_type = row.get('ticket_type', 'ALL') if 'ticket_type' in df.columns else ticket_type
                ticket_type = str(ticket_type).strip() if pd.notna(ticket_type) else ticket_type

                # 只添加有产品名称的数据
                if product_name and product_name != 'nan':
                    product = AirlineProduct(
                        airline=airline,
                        product_name=product_name,
                        route=route,
                        booking_class=booking_class,
                        price_increase=price_increase,
                        rebate_ratio=rebate_ratio,
                        office=office,
                        remarks=remarks,
                        valid_period=valid_period,
                        ticket_type=ticket_type,
                        policy_identifier=policy_identifier
                    )
                    products.append(product)

            except Exception as e:
                print(f"  [X] 第 {index} 行处理失败: {e}")
                continue

    return products


# ============================================================================
# 全局智能体实例 (用于工具函数访问)
# ============================================================================

# 延迟初始化的全局智能体实例
_agent_instance = None

def get_agent_instance(reload=False):
    """
    获取或初始化智能体实例（每次都重新加载数据以确保数据最新）

    Args:
        reload: 是否强制重新加载数据（已废弃，现在总是重新加载）

    Returns:
        IntelligentAgent: 智能体实例
    """
    # 获取脚本所在目录的父目录作为workspace路径
    workspace_path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    # 每次都创建新实例，确保数据最新
    _agent_instance = IntelligentAgent()

    # 初始化智能体
    try:
        df = load_cleaned_data()
        if not df.empty:
            products = create_airline_products(df)
            _agent_instance.add_products(products)
            _agent_instance.load_documentation()
            print(f"\n[OK] 智能体初始化完成，已加载 {len(products)} 个产品")

            # 自动更新航司为空的产品
            update_null_airlines(_agent_instance)
        else:
            print("[WARN] 未找到数据，智能体将在空数据状态下运行")
    except Exception as e:
        print(f"[WARN] 智能体初始化失败: {e}")

    return _agent_instance


# ============================================================================
# LangChain 工具定义
# ============================================================================

@tool
def search_airline_products(keyword: str, ticket_type: str = None, runtime: ToolRuntime = None) -> str:
    """
    搜索航空公司产品信息

    Args:
        keyword: 搜索关键词（产品名称、航线、备注等）
        ticket_type: 可选，票证类型筛选（GP/BSP/B2B/MULTI/ALL）

    Returns:
        匹配的产品列表（JSON格式）
    """
    ctx = runtime.context if runtime else new_context(method="search_airline_products")
    agent = get_agent_instance()

    # 处理票证类型参数
    ticket_type_upper = ticket_type.strip().upper() if ticket_type and ticket_type.strip() else None
    valid_types = ['GP', 'BSP', 'B2B', 'MULTI', 'ALL']

    if ticket_type_upper and ticket_type_upper not in valid_types:
        return json.dumps({
            "message": f"无效的票证类型: {ticket_type}。支持的类型: {', '.join(valid_types)}",
            "count": 0
        }, ensure_ascii=False, indent=2)

    results = agent.search_products(keyword, ticket_type_upper)

    if not results:
        return json.dumps({"message": "未找到匹配的产品", "count": 0}, ensure_ascii=False, indent=2)

    products_data = [p.to_dict() for p in results[:10]]  # 限制返回10条
    return json.dumps({
        "message": f"找到 {len(results)} 个匹配产品（显示前10条）",
        "count": len(results),
        "products": products_data
    }, ensure_ascii=False, indent=2)


@tool
def get_products_by_airline(airline_code: str, runtime: ToolRuntime = None) -> str:
    """
    根据航空公司代码查询产品

    Args:
        airline_code: 航空公司代码（如：MU, CZ, HU, CA等）

    Returns:
        该航空公司的所有产品列表
    """
    ctx = runtime.context if runtime else new_context(method="get_products_by_airline")
    agent = get_agent_instance()
    results = agent.get_products_by_airline(airline_code.strip().upper())

    if not results:
        return json.dumps({"message": f"未找到航空公司 {airline_code} 的产品", "count": 0}, ensure_ascii=False, indent=2)

    products_data = [p.to_dict() for p in results]
    return json.dumps({
        "message": f"航空公司 {airline_code} 共有 {len(results)} 个产品",
        "count": len(results),
        "products": products_data
    }, ensure_ascii=False, indent=2)


@tool
def get_products_by_ticket_type(ticket_type: str, runtime: ToolRuntime = None) -> str:
    """
    根据票证类型查询产品

    Args:
        ticket_type: 票证类型（GP/BSP/B2B/MULTI/ALL）
            - GP: 集团客户产品
            - BSP: 中性票产品
            - B2B: 企业对企业产品
            - MULTI: 支持多种票证类型的产品
            - ALL: 通用型产品

    Returns:
        该票证类型的产品列表
    """
    ctx = runtime.context if runtime else new_context(method="get_products_by_ticket_type")
    agent = get_agent_instance()

    # 验证票证类型
    valid_types = ['GP', 'BSP', 'B2B', 'MULTI', 'ALL']
    ticket_type_upper = ticket_type.strip().upper()

    if ticket_type_upper not in valid_types:
        return json.dumps({
            "message": f"无效的票证类型: {ticket_type}。支持的类型: {', '.join(valid_types)}",
            "count": 0
        }, ensure_ascii=False, indent=2)

    results = agent.get_products_by_ticket_type(ticket_type_upper)

    if not results:
        return json.dumps({"message": f"未找到票证类型为 {ticket_type_upper} 的产品", "count": 0}, ensure_ascii=False, indent=2)

    products_data = [p.to_dict() for p in results]
    return json.dumps({
        "message": f"票证类型 {ticket_type_upper} 共有 {len(results)} 个产品",
        "count": len(results),
        "products": products_data
    }, ensure_ascii=False, indent=2)


@tool
def get_all_airlines(runtime: ToolRuntime = None) -> str:
    """
    获取所有航空公司列表

    Returns:
        所有航空公司代码和名称列表
    """
    ctx = runtime.context if runtime else new_context(method="get_all_airlines")
    agent = get_agent_instance()

    # 获取航司代码和名称的映射
    airline_info = {}

    # 从产品中提取航司代码
    for product in agent.products:
        if product.airline not in airline_info:
            airline_info[product.airline] = product.airline  # 默认使用代码

    # 尝试从数据中获取航司名称
    workspace_path = os.getenv("COZE_WORKSPACE_PATH", "/workspace/projects")
    all_data_file = os.path.join(workspace_path, "assets", "airline_products_all.csv")

    if os.path.exists(all_data_file):
        try:
            df = pd.read_csv(all_data_file, encoding='utf-8-sig')
            if 'airline_code' in df.columns and 'airline_name' in df.columns:
                for _, row in df.iterrows():
                    code = row['airline_code']
                    name = row['airline_name']
                    # 使用标量值的非空检查
                    if code is not None and name is not None and code != '' and name != '':
                        code_str = str(code)
                        name_str = str(name)
                        airline_info[code_str] = name_str
        except Exception as e:
            print(f"[WARN] 加载航司名称失败: {e}")

    if not airline_info:
        return json.dumps({"message": "当前没有可用航空公司数据", "airlines": []}, ensure_ascii=False, indent=2)

    airlines_list = [{"code": code, "name": name} for code, name in sorted(airline_info.items())]

    return json.dumps({
        "message": f"共有 {len(airlines_list)} 个航空公司",
        "count": len(airlines_list),
        "airlines": airlines_list
    }, ensure_ascii=False, indent=2)


@tool
def get_agent_statistics(runtime: ToolRuntime = None) -> str:
    """
    获取智能体统计信息

    Returns:
        统计信息（产品总数、航空公司数量、航线数量等）
    """
    ctx = runtime.context if runtime else new_context(method="get_agent_statistics")
    agent = get_agent_instance()
    stats = agent.get_statistics()

    return json.dumps({
        "message": "智能体统计信息",
        "statistics": stats
    }, ensure_ascii=False, indent=2)


@tool
def get_documentation_help(topic: str = None, runtime: ToolRuntime = None) -> str:
    """
    获取文档帮助信息

    Args:
        topic: 帮助主题（可选），不提供则返回总体帮助

    Returns:
        帮助信息内容
    """
    ctx = runtime.context if runtime else new_context(method="get_documentation_help")
    agent = get_agent_instance()

    if not agent.documentation:
        return "文档系统未加载"

    if topic:
        results = agent.documentation.search_all_documents(topic)
        if results:
            help_text = f"找到 '{topic}' 的相关帮助信息：\n\n"
            for doc_key, matches in results.items():
                help_text += f"【文档：{doc_key}】\n"
                for match in matches[:3]:  # 只显示前3个匹配
                    help_text += f"  行 {match['line_number']}: {match['content'][:100]}\n"
                help_text += "\n"
            return help_text
        else:
            return f"未找到关于 '{topic}' 的帮助信息"
    else:
        # 获取快速开始指南
        quickstart = agent.documentation.get_document('quickstart')
        if quickstart:
            return quickstart[:1500] + "\n\n(更多内容请查看完整文档)"
        else:
            return "未找到快速开始文档"


@tool
def export_products_to_excel(runtime: ToolRuntime = None) -> str:
    """
    导出所有产品数据到Excel文件

    Returns:
        导出结果和文件路径
    """
    ctx = runtime.context if runtime else new_context(method="export_products_to_excel")
    agent = get_agent_instance()

    workspace_path = os.getenv("COZE_WORKSPACE_PATH", "/workspace/projects")
    output_dir = os.path.join(workspace_path, "assets", "agent_data_with_docs")
    os.makedirs(output_dir, exist_ok=True)

    output_file = os.path.join(output_dir, "所有产品数据_export.xlsx")

    try:
        agent.export_to_excel(output_file)
        return json.dumps({
            "message": "导出成功",
            "file_path": output_file
        }, ensure_ascii=False, indent=2)
    except Exception as e:
        return json.dumps({
            "message": f"导出失败: {str(e)}"
        }, ensure_ascii=False, indent=2)


# ============================================================================
# LangChain Agent 定义
# ============================================================================

LLM_CONFIG = "config/agent_llm_config.json"

# 默认保留最近 20 轮对话 (40 条消息)
MAX_MESSAGES = 40

def _windowed_messages(old, new):
    """滑动窗口: 只保留最近 MAX_MESSAGES 条消息"""
    return add_messages(old, new)[-MAX_MESSAGES:]  # type: ignore


class AgentState(MessagesState):
    messages: Annotated[list[AnyMessage], _windowed_messages]


def build_agent(ctx=None):
    """
    构建航空公司产品智能体

    Returns:
        LangChain Agent实例
    """
    workspace_path = os.getenv("COZE_WORKSPACE_PATH", "/workspace/projects")
    config_path = os.path.join(workspace_path, LLM_CONFIG)

    # 读取配置文件
    with open(config_path, 'r', encoding='utf-8') as f:
        cfg = json.load(f)

    # 配置LLM
    api_key = os.getenv("COZE_WORKLOAD_IDENTITY_API_KEY")
    base_url = os.getenv("COZE_INTEGRATION_MODEL_BASE_URL")

    llm = ChatOpenAI(
        model=cfg['config'].get("model"),
        api_key=api_key,
        base_url=base_url,
        temperature=cfg['config'].get('temperature', 0.7),
        streaming=True,
        timeout=cfg['config'].get('timeout', 600),
        extra_body={
            "thinking": {
                "type": cfg['config'].get('thinking', 'disabled')
            }
        },
        default_headers=default_headers(ctx) if ctx else {}
    )

    # 定义工具列表
    tools = [
        search_airline_products,
        get_products_by_airline,
        get_all_airlines,
        get_agent_statistics,
        get_documentation_help,
        export_products_to_excel
    ]

    # 创建Agent
    agent = create_agent(
        model=llm,
        system_prompt=cfg.get("sp"),
        tools=tools,
        checkpointer=get_memory_saver(),
        state_schema=AgentState,
    )

    return agent


# ============================================================================
# 主函数 (用于直接运行测试)
# ============================================================================

def main():
    """主函数"""
    print("="*80)
    print("航空公司产品智能体系统 (集成文档系统 + LangChain)")
    print("="*80)
    print()

    # 显示统计信息
    agent = get_agent_instance()
    agent.display_summary()

    # 提供交互式查询
    print("\n智能体已就绪，你可以通过以下工具进行操作：")
    print("  1. search_airline_products - 搜索产品")
    print("  2. get_products_by_airline - 按航空公司查询")
    print("  3. get_all_airlines - 获取所有航空公司")
    print("  4. get_agent_statistics - 获取统计信息")
    print("  5. get_documentation_help - 获取帮助")
    print("  6. export_products_to_excel - 导出数据")
    print()


if __name__ == "__main__":
    main()
